#!/usr/bin/env python3
"""Applique les notices issues du pipeline ISBN au classeur avant génération du site."""

import json
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "inventaire_bibliotheque.xlsx"
IMPORTS = ROOT / "data" / "bibliotheque_pipeline.json"


def clean(value):
    return "" if value is None else str(value).strip()


def main():
    if not IMPORTS.exists():
        return 0
    wb = load_workbook(WORKBOOK)
    ws = wb["Catalogue"]
    headers = [clean(cell.value) for cell in ws[1]]
    by_name = {name: i + 1 for i, name in enumerate(headers) if name}
    existing = {
        clean(ws.cell(row=r, column=by_name["ID"]).value).upper(): r
        for r in range(2, ws.max_row + 1)
        if clean(ws.cell(row=r, column=by_name["ID"]).value)
    }
    records = json.loads(IMPORTS.read_text(encoding="utf-8"))
    for record in records:
        book_id = clean(record.get("ID")).upper()
        if not book_id:
            continue
        row = existing.get(book_id, ws.max_row + 1)
        for key, value in record.items():
            column = by_name.get(key)
            if column:
                ws.cell(row=row, column=column, value=value)
        existing[book_id] = row
    wb.save(WORKBOOK)
    print(f"{len(records)} notice(s) du pipeline appliquée(s) au classeur.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
