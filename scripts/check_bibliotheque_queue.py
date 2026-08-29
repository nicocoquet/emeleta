#!/usr/bin/env python3
"""Contrôle la file d’ingestion ISBN et propose le prochain identifiant BIB."""

from __future__ import annotations

import json
import re
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "inventaire_bibliotheque.xlsx"
PIPELINE = ROOT / "data" / "bibliotheque_pipeline.json"
QUEUE = ROOT / "photos" / "bibliotheque" / "isbn" / "a_traiter"
VERIFY = ROOT / "photos" / "bibliotheque" / "isbn" / "a_verifier"
DONE = ROOT / "photos" / "bibliotheque" / "isbn" / "traite"
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".heic", ".heif"}


def clean(value) -> str:
    return "" if value is None else str(value).strip()


def bib_number(value: str) -> int | None:
    match = re.fullmatch(r"BIB-(\d+)", clean(value).upper())
    return int(match.group(1)) if match else None


def existing_ids() -> set[int]:
    ids: set[int] = set()
    if WORKBOOK.exists():
        wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
        if "Catalogue" in wb.sheetnames:
            ws = wb["Catalogue"]
            headers = [clean(c.value) for c in ws[1]]
            if "ID" in headers:
                col = headers.index("ID") + 1
                for row in range(2, ws.max_row + 1):
                    number = bib_number(ws.cell(row=row, column=col).value)
                    if number is not None:
                        ids.add(number)
    if PIPELINE.exists():
        for record in json.loads(PIPELINE.read_text(encoding="utf-8")):
            number = bib_number(record.get("ID"))
            if number is not None:
                ids.add(number)
    return ids


def images(path: Path) -> list[Path]:
    if not path.exists():
        return []
    return sorted(
        p for p in path.iterdir()
        if p.is_file() and p.suffix.lower() in IMAGE_EXTENSIONS
    )


def main() -> int:
    ids = existing_ids()
    next_number = max(ids, default=0) + 1
    queued = images(QUEUE)
    verify = images(VERIFY)
    done = images(DONE)

    print(f"Prochain identifiant disponible : BIB-{next_number:03d}")
    print(f"À traiter : {len(queued)} image(s)")
    for item in queued:
        print(f"  - {item.name}")
    print(f"À vérifier : {len(verify)} image(s)")
    print(f"Traitées : {len(done)} image(s)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
