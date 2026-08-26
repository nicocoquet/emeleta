#!/usr/bin/env python3
"""Génère la bibliothèque MkDocs depuis inventaire_bibliotheque.xlsx."""

from __future__ import annotations

import html
import re
import shutil
import sys
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "inventaire_bibliotheque.xlsx"
SOURCE_PHOTOS = ROOT / "photos" / "bibliotheque"
DOCS = ROOT / "docs"
LIBRARY = DOCS / "bibliotheque"
SITE_IMAGES = DOCS / "assets" / "images" / "bibliotheque"

# Facettes retenues pour la bibliothèque Emeleta : pas de localisation,
# car l'immense majorité des ouvrages est conservée au Bureau.
FACETS = [
    ("author", "Auteur / éditeur scientifique"),
    ("publication_date", "Date de publication"),
    ("document_type", "Type de document"),
    ("subject", "Thème"),
    ("language", "Langue"),
]

# Colonnes nécessaires au fonctionnement du site. Les colonnes d'enrichissement
# ISBN / notices sont optionnelles afin de ne pas bloquer la publication si elles
# sont progressivement ajoutées au tableur.
REQUIRED_COLUMNS = {
    "ID", "Auteur_editeur_scientifique", "Titre", "Date_publication", "Langue",
    "Type_document", "Sujet", "Mots_cles", "Serie_liee",
    "Photo_couverture", "Photo_bibliographique",
}


def text(value) -> str:
    return "" if value is None else str(value).strip()


def first_value(book: dict, *names: str) -> str:
    for name in names:
        value = text(book.get(name))
        if value:
            return value
    return ""


def split_values(value) -> list[str]:
    return [item.strip() for item in text(value).split(";") if item.strip()]


def normalize_author(value: str) -> str:
    return re.sub(r"\s*\[[^\]]+\]\s*$", "", value).strip()


def display_author(value: str) -> str:
    authors = []
    for item in split_values(value):
        match = re.match(r"^(.*?)\s*\[([^\]]+)\]\s*$", item)
        if match:
            authors.append(f"{match.group(1).strip()} ({match.group(2).strip()})")
        else:
            authors.append(item)
    return " · ".join(authors)


def display_responsibilities(value: str) -> str:
    items = []
    for item in split_values(value):
        match = re.match(r"^(.*?)\s*\[([^\]]+)\]\s*$", item)
        if match:
            items.append(f"{match.group(1).strip()} ({match.group(2).strip()})")
        else:
            items.append(item)
    return " · ".join(items)


def author_facets(value: str) -> list[str]:
    return [normalize_author(item) for item in split_values(value)]


def publication_bucket(value) -> str:
    match = re.search(r"\b(\d{4})\b", text(value))
    if not match:
        return "Date à préciser"
    year = int(match.group(1))
    if year < 1900:
        return "Avant 1900"
    if year <= 1945:
        return "1900–1945"
    if year <= 1980:
        return "1946–1980"
    if year <= 2000:
        return "1981–2000"
    return "Après 2000"


def slug_sort(value: str) -> str:
    normalized = unicodedata.normalize("NFD", value)
    return "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn").casefold()


def rows_as_dicts(sheet):
    headers = [text(cell.value) for cell in sheet[1]]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(value not in (None, "") for value in row):
            continue
        yield {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}


def require_columns(sheet):
    actual = {text(cell.value) for cell in sheet[1]}
    missing = REQUIRED_COLUMNS - actual
    if missing:
        raise ValueError("Colonnes manquantes dans Catalogue : " + ", ".join(sorted(missing)))


def data_attr(values) -> str:
    return "||".join(values)


def copy_photo(filename: str, warnings: list[str]) -> str:
    filename = Path(filename).name
    if not filename:
        return ""
    source = SOURCE_PHOTOS / filename
    if not source.is_file():
        warnings.append(f"Photographie introuvable : {filename}")
        return ""
    SITE_IMAGES.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, SITE_IMAGES / filename)
    return filename


def field(label: str, value: str) -> str:
    if not value:
        return ""
    return f'<div class="record-field"><dt>{html.escape(label)}</dt><dd>{html.escape(value)}</dd></div>'


def metadata_html(book: dict) -> str:
    publication = ", ".join(
        part for part in [
            text(book.get("Lieu_publication")),
            text(book.get("Editeur")),
            text(book.get("Date_publication")),
        ] if part
    )
    fields = [
        ("Auteur / éditeur scientifique", display_author(text(book.get("Auteur_editeur_scientifique")))),
        ("Autre responsabilité", display_responsibilities(first_value(book, "Autre_responsabilite", "Autre_responsabilité"))),
        ("Sous-titre", text(book.get("Sous_titre"))),
        ("Volume", text(book.get("Volume"))),
        ("Publication", publication),
        ("Édition", text(book.get("Edition"))),
        ("Collection", text(book.get("Collection"))),
        ("Numéro dans la collection", first_value(book, "Num_collection", "Numero_collection")),
        ("ISBN", first_value(book, "ISBN", "ISBN_13")),
        ("ISBN-13", first_value(book, "ISBN_13", "EAN")),
        ("ISBN-10", text(book.get("ISBN_10"))),
        ("Langue", text(book.get("Langue"))),
        ("Langue originale", text(book.get("Langue_originale"))),
        ("Titre original", text(book.get("Titre_original"))),
        ("Type de document", text(book.get("Type_document"))),
        ("Genre", text(book.get("Genre"))),
        ("Nombre de pages", text(book.get("Nbr_pages"))),
        ("Dimensions", text(book.get("Dimensions"))),
        ("Localisation", text(book.get("Localisation"))),
    ]
    return "\n".join(field(label, value) for label, value in fields if value)


def indexing_html(book: dict) -> str:
    fields = [
        ("Thème", text(book.get("Sujet"))),
        ("Lieu", text(book.get("Lieu_sujet"))),
        ("Période", text(book.get("Periode_sujet"))),
        ("Mots-clés", text(book.get("Mots_cles"))),
    ]
    return "\n".join(field(label, value) for label, value in fields if value)


def notice_html(book: dict) -> str:
    fields = [
        ("Source de la notice", text(book.get("Source_notice"))),
        ("Identifiant de la notice", text(book.get("Identifiant_notice"))),
        ("Statut de la notice", text(book.get("Statut_notice"))),
        ("Confiance", text(book.get("Confiance_notice"))),
        ("Notes de contrôle", text(book.get("Notes_controle"))),
    ]
    return "\n".join(field(label, value) for label, value in fields if value)


def exemplar_html(book: dict) -> str:
    fields = [
        ("Ex-libris", text(book.get("Ex_libris"))),
        ("Dédicace / annotations", text(book.get("Dedicace_annotations"))),
        ("Notes", text(book.get("Notes"))),
    ]
    return "\n".join(field(label, value) for label, value in fields if value)


def render_select(name: str, label: str, values: list[str]) -> str:
    options = ['<option value="">Tous</option>']
    options.extend(
        f'<option value="{html.escape(value, quote=True)}">{html.escape(value)}</option>'
        for value in sorted(values, key=slug_sort)
    )
    return (
        f'<div class="catalog-filter"><label for="filter-{name}">{html.escape(label)}</label>'
        f'<select id="filter-{name}" name="{name}" data-filter="{name}">' + "".join(options) + "</select></div>"
    )


def main() -> int:
    if not WORKBOOK.exists():
        print(f"Classeur introuvable : {WORKBOOK}", file=sys.stderr)
        return 1
    wb = load_workbook(WORKBOOK, data_only=True)
    if "Catalogue" not in wb.sheetnames:
        print("Feuille Catalogue introuvable.", file=sys.stderr)
        return 1
    sheet = wb["Catalogue"]
    require_columns(sheet)
    books = [row for row in rows_as_dicts(sheet) if text(row.get("ID")) and text(row.get("Titre"))]
    known_ids = {text(book["ID"]).upper() for book in books}
    LIBRARY.mkdir(parents=True, exist_ok=True)
    SITE_IMAGES.mkdir(parents=True, exist_ok=True)
    for legacy in LIBRARY.glob("BIB-*.fr.md"):
        legacy.unlink()
    facet_values = {name: set() for name, _ in FACETS}
    cards = []
    warnings: list[str] = []
    for book in books:
        book_id = text(book["ID"]).upper()
        title = text(book["Titre"])
        authors = author_facets(text(book.get("Auteur_editeur_scientifique")))
        subjects = split_values(book.get("Sujet"))
        publication_group = publication_bucket(book.get("Date_publication"))
        languages = split_values(book.get("Langue"))
        document_types = split_values(book.get("Type_document"))
        per_book_facets = {
            "author": authors,
            "publication_date": [publication_group],
            "document_type": document_types,
            "subject": subjects,
            "language": languages,
        }
        for name, values in per_book_facets.items():
            facet_values[name].update(values)
        cover = copy_photo(text(book.get("Photo_couverture")), warnings)
        biblio_photo = copy_photo(text(book.get("Photo_bibliographique")), warnings)
        extra_photos = [
            copied for item in split_values(book.get("Photos_supplementaires"))
            if (copied := copy_photo(item, warnings))
        ]
        cover_html = (
            f'<a href="../../assets/images/bibliotheque/{html.escape(cover)}" target="_blank">'
            f'<img src="../../assets/images/bibliotheque/{html.escape(cover)}" alt="Couverture — {html.escape(title)}"></a>'
            if cover else '<div class="record-placeholder book-placeholder">Sans photographie</div>'
        )
        gallery_items = []
        if biblio_photo:
            gallery_items.append(
                f'<figure><a href="../../assets/images/bibliotheque/{html.escape(biblio_photo)}" target="_blank">'
                f'<img src="../../assets/images/bibliotheque/{html.escape(biblio_photo)}" alt="Informations bibliographiques — {html.escape(title)}" loading="lazy"></a>'
                f'<figcaption>Page de titre / informations bibliographiques</figcaption></figure>'
            )
        for filename in extra_photos:
            gallery_items.append(
                f'<figure><a href="../../assets/images/bibliotheque/{html.escape(filename)}" target="_blank">'
                f'<img src="../../assets/images/bibliotheque/{html.escape(filename)}" alt="{html.escape(title)}" loading="lazy"></a>'
                f'<figcaption>Photographie complémentaire</figcaption></figure>'
            )
        gallery = "\n".join(gallery_items) or '<p class="empty-state">Aucune photographie complémentaire.</p>'
        links = []
        for associated in re.findall(r"BIB-\d{3,}", text(book.get("Serie_liee")), flags=re.IGNORECASE):
            associated = associated.upper()
            if associated == book_id:
                continue
            if associated in known_ids:
                links.append(f'<a href="../{associated}/">{html.escape(associated)}</a>')
            else:
                warnings.append(f"{book_id} : volume lié inconnu : {associated}")
        related = " · ".join(links) or "_Aucun volume associé._"
        external = first_value(book, "URL_notice", "URL_source")
        external_block = f'[{html.escape(external)}]({external})' if external else "_Aucune notice externe renseignée._"
        notice = notice_html(book)
        page = f'''<div class="lot-hero book-hero">\n<div class="lot-visual book-cover">{cover_html}</div>\n<div class="lot-summary">\n<p class="record-kicker">Bibliothèque · {html.escape(book_id)}</p>\n<h1>{html.escape(title)}</h1>\n<p class="lot-dating">{html.escape(display_author(text(book.get("Auteur_editeur_scientifique"))))}</p>\n<p class="lot-estimate">{html.escape(", ".join(part for part in [text(book.get("Lieu_publication")), text(book.get("Editeur")), text(book.get("Date_publication"))] if part))}</p>\n</div>\n</div>\n\n## Notice bibliographique\n\n<dl class="record-metadata">\n{metadata_html(book)}\n</dl>\n\n## Indexation\n\n<dl class="record-metadata">\n{indexing_html(book)}\n</dl>\n\n## Données de notice\n\n<dl class="record-metadata">\n{notice or '<div class="record-field"><dd>Aucune donnée de notice renseignée.</dd></div>'}\n</dl>\n\n## Exemplaire d'Emeleta\n\n<dl class="record-metadata">\n{exemplar_html(book) or '<div class="record-field"><dd>Aucune particularité renseignée.</dd></div>'}\n</dl>\n\n## Volumes associés\n\n{related}\n\n## Notice externe\n\n{external_block}\n\n## Photographies\n\n<div class="photo-grid book-photo-grid">\n{gallery}\n</div>\n\n[← Retour à la bibliothèque](../)\n'''
        (LIBRARY / f"{book_id}.fr.md").write_text(page, encoding="utf-8")
        attrs = " ".join(
            f'data-{name.replace("_", "-" )}="{html.escape(data_attr(values), quote=True)}"'
            for name, values in per_book_facets.items()
        )
        visual = (
            f'<img src="../assets/images/bibliotheque/{html.escape(cover)}" alt="{html.escape(title)}" loading="lazy">'
            if cover else '<div class="card-placeholder book-placeholder">Sans photographie</div>'
        )
        author_label = " · ".join(authors) or "Auteur non renseigné"
        pub_label = ", ".join(part for part in [text(book.get("Editeur")), text(book.get("Date_publication"))] if part)
        card_text = " | ".join(
            part for part in [title, author_label, pub_label, text(book.get("Mots_cles")), text(book.get("ISBN")), text(book.get("ISBN_13"))]
            if part
        )
        cards.append(
            f'<article class="catalog-card book-card" {attrs} data-search="{html.escape(card_text, quote=True)}">'
            f'<a href="{html.escape(book_id)}/">{visual}</a>'
            f'<div class="catalog-card-body"><div class="card-lot"><span>{html.escape(book_id)}</span><span>{html.escape(text(book.get("Type_document")))}</span></div>'
            f'<h2><a href="{html.escape(book_id)}/">{html.escape(title)}</a></h2>'
            f'<p class="book-author">{html.escape(author_label)}</p><p class="card-date">{html.escape(pub_label)}</p></div></article>'
        )
    selects = "\n".join(render_select(name, label, list(facet_values[name])) for name, label in FACETS)
    index = f'''<div class="catalogue-heading">\n<p class="eyebrow">Emeleta · Bibliothèque</p>\n<h1>Bibliothèque</h1>\n<p>Catalogue des ouvrages conservés à Emeleta. Les filtres peuvent être combinés.</p>\n</div>\n\n<form class="catalog-filters library-filters" data-catalog-filters>\n<div class="catalog-filter catalog-filter-search"><label for="catalog-search">Recherche</label><input id="catalog-search" type="search" name="q" placeholder="Titre, auteur, ISBN, mot-clé…" autocomplete="off"></div>\n{selects}\n<button class="catalog-reset" type="reset">Réinitialiser</button>\n<p class="catalog-result"><strong data-result-count>{len(cards)}</strong> ouvrage(s)</p>\n</form>\n\n<div class="catalog-grid book-grid" data-catalog-grid>\n{chr(10).join(cards)}\n</div>\n\n<p class="catalog-empty" data-catalog-empty hidden>Aucun ouvrage ne correspond à ces critères.</p>\n'''
    (LIBRARY / "index.fr.md").write_text(index, encoding="utf-8")
    print(f"{len(cards)} ouvrage(s) généré(s).")
    if warnings:
        print("Avertissements :", file=sys.stderr)
        for warning in sorted(set(warnings)):
            print(f"- {warning}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
