#!/usr/bin/env python3
"""Génère les pages MkDocs depuis inventaire_mobilier.xlsx."""

from __future__ import annotations

import html
import re
import shutil
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "inventaire_mobilier.xlsx"
SOURCE_PHOTOS = ROOT / "photos"
DOCS = ROOT / "docs"
CATALOG = DOCS / "catalogue"
SITE_IMAGES = DOCS / "assets" / "images"


def text(value) -> str:
    return "" if value is None else str(value).strip()


def slug(value: str) -> str:
    value = value.lower().strip()
    value = re.sub(r"[^a-z0-9-]+", "-", value)
    return value.strip("-")


def money(value) -> str:
    if value in (None, ""):
        return ""
    try:
        return f"{float(value):,.0f} €".replace(",", " ")
    except (TypeError, ValueError):
        return text(value)


def rows_as_dicts(sheet, header_row: int = 3):
    headers = [text(cell.value) for cell in sheet[header_row]]
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        values = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        if any(value not in (None, "") for value in row):
            yield values


def require_columns(actual: set[str], expected: set[str], sheet_name: str):
    missing = expected - actual
    if missing:
        raise ValueError(f"Feuille {sheet_name}: colonnes manquantes: {', '.join(sorted(missing))}")


def main() -> int:
    if not WORKBOOK.exists():
        print(f"Classeur introuvable: {WORKBOOK}", file=sys.stderr)
        return 1

    wb = load_workbook(WORKBOOK, data_only=True)
    require_columns(
        {text(c.value) for c in wb["Mobilier"][3]},
        {"Identifiant", "Titre", "Catégorie", "Description", "Publié"},
        "Mobilier",
    )
    require_columns(
        {text(c.value) for c in wb["Photos"][3]},
        {"Identifiant objet", "Fichier", "Ordre", "Légende", "Image principale"},
        "Photos",
    )

    furniture = [r for r in rows_as_dicts(wb["Mobilier"]) if text(r["Publié"]).lower() in {"oui", "yes", "true", "1"}]
    photo_rows = list(rows_as_dicts(wb["Photos"]))
    photos_by_object = defaultdict(list)
    for photo in photo_rows:
        photos_by_object[text(photo["Identifiant objet"])].append(photo)
    for items in photos_by_object.values():
        items.sort(key=lambda p: (p["Ordre"] if isinstance(p["Ordre"], (int, float)) else 999, text(p["Fichier"])))

    CATALOG.mkdir(parents=True, exist_ok=True)
    SITE_IMAGES.mkdir(parents=True, exist_ok=True)
    for old in CATALOG.glob("MOB-*.md"):
        old.unlink()

    cards = []
    categories = defaultdict(int)
    warnings = []

    for obj in furniture:
        object_id = text(obj["Identifiant"])
        title = text(obj["Titre"])
        if not object_id or not title:
            warnings.append("Une ligne publiée sans identifiant ou sans titre a été ignorée.")
            continue

        photos = photos_by_object.get(object_id, [])
        available = []
        for photo in photos:
            filename = Path(text(photo["Fichier"])).name
            source = SOURCE_PHOTOS / filename
            if source.exists():
                shutil.copy2(source, SITE_IMAGES / filename)
                available.append(photo)
            else:
                warnings.append(f"{object_id}: photographie introuvable: {filename}")

        cover = next((p for p in available if text(p["Image principale"]).lower() in {"oui", "yes", "true", "1"}), available[0] if available else None)
        category = text(obj["Catégorie"]) or "Non classé"
        categories[category] += 1

        low = money(obj.get("Estimation basse (€)"))
        high = money(obj.get("Estimation haute (€)"))
        estimate = " – ".join(v for v in [low, high] if v) or "Non renseignée"

        metadata = [
            ("Identifiant", object_id),
            ("Catégorie", category),
            ("Datation", text(obj.get("Datation"))),
            ("Origine", text(obj.get("Origine"))),
            ("Matériaux", text(obj.get("Matériaux"))),
            ("Dimensions", text(obj.get("Dimensions"))),
            ("Localisation", text(obj.get("Localisation"))),
            ("Statut", text(obj.get("Statut"))),
            ("Estimation indicative", estimate),
        ]
        meta_html = "\n".join(
            f'<div class="record-field"><dt>{html.escape(label)}</dt><dd>{html.escape(value)}</dd></div>'
            for label, value in metadata if value
        )

        gallery = []
        for photo in available:
            filename = Path(text(photo["Fichier"])).name
            caption = text(photo["Légende"]) or title
            gallery.append(
                f'<figure><a href="../../assets/images/{html.escape(filename)}" target="_blank">'
                f'<img src="../../assets/images/{html.escape(filename)}" alt="{html.escape(caption)}" loading="lazy"></a>'
                f'<figcaption>{html.escape(caption)}</figcaption></figure>'
            )
        gallery_html = "\n".join(gallery) if gallery else '<p class="empty-state">Aucune photographie disponible.</p>'

        sources = []
        for key in ["Source 1", "Source 2"]:
            url = text(obj.get(key))
            if url:
                sources.append(f"- [{html.escape(url)}]({url})")
        source_block = "\n".join(sources) if sources else "_Aucune source renseignée._"

        page = f"""# {title}

<p class="record-kicker">{html.escape(object_id)} · {html.escape(category)}</p>

<dl class="record-grid">
{meta_html}
</dl>

## Description

{text(obj.get("Description")) or "_Description à compléter._"}

## État de conservation

{text(obj.get("État")) or "_État à compléter._"}

## Photographies

<div class="gallery">
{gallery_html}
</div>

## Sources et comparaisons

{source_block}

<p class="notice">L’identification et l’estimation sont indicatives et peuvent évoluer avec de nouvelles mesures, photographies ou expertises.</p>
"""
        (CATALOG / f"{object_id}.md").write_text(page, encoding="utf-8")

        cover_html = (
            f'<img src="assets/images/{html.escape(Path(text(cover["Fichier"])).name)}" alt="{html.escape(title)}">'
            if cover else '<div class="card-placeholder">Sans photographie</div>'
        )
        cards.append(
            f'<article class="catalog-card"><a href="catalogue/{object_id}/">{cover_html}'
            f'<div class="catalog-card-body"><span>{html.escape(category)}</span><h2>{html.escape(title)}</h2>'
            f'<p>{html.escape(text(obj.get("Datation")) or "Datation à préciser")}</p></div></a></article>'
        )

    category_text = " · ".join(f"{html.escape(k)} ({v})" for k, v in sorted(categories.items()))
    index = f"""# Inventaire du mobilier

<div class="hero-panel">
  <p class="eyebrow">Catalogue illustré</p>
  <h2>{len(cards)} objet{'s' if len(cards) != 1 else ''} documenté{'s' if len(cards) != 1 else ''}</h2>
  <p>Un inventaire évolutif du mobilier, de ses caractéristiques matérielles et de son état de conservation.</p>
</div>

<p class="category-line">{category_text}</p>

<div class="catalog-grid">
{''.join(cards) if cards else '<p class="empty-state">Aucun objet publié.</p>'}
</div>
"""
    (DOCS / "index.md").write_text(index, encoding="utf-8")
    (CATALOG / "index.md").write_text(
        "# Catalogue\n\nToutes les fiches publiées apparaissent ci-dessous. Utilisez la recherche en haut de la page pour retrouver un objet, une matière, une époque ou une localisation.\n\n"
        + "\n".join(f"- [{text(o['Titre'])}]({text(o['Identifiant'])}.md) — {text(o['Datation'])}" for o in furniture),
        encoding="utf-8",
    )

    if warnings:
        print("Avertissements:")
        for warning in warnings:
            print(f"- {warning}")
    print(f"Catalogue généré: {len(cards)} objet(s), {sum(len(v) for v in photos_by_object.values())} photographie(s) référencée(s).")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

