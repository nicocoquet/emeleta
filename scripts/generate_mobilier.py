#!/usr/bin/env python3
"""Construit la partie « Mobilier » du site MkDocs à partir d'Excel.

Entrées : ``inventaire_mobilier.xlsx`` (feuilles ``Mobilier`` et ``Photos``)
et les originaux du dossier ``photos/``.

Sorties : fiches et index sous ``docs/inventaire/``, copies d'images sous
``docs/assets/images/`` et page Statistiques commune aux deux inventaires.
Cette page contient le mobilier ainsi que le conteneur dont les données seront
créées ensuite par ``generate_bibliotheque.py``.

Le script est déterministe et peut donc être relancé à chaque déploiement.
"""

from __future__ import annotations

import html
import re
import shutil
import sys
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "inventaire_mobilier.xlsx"
SOURCE_PHOTOS = ROOT / "photos"
DOCS = ROOT / "docs"
INVENTORY_DIR = DOCS / "inventaire"
LEGACY_CATALOG_DIR = DOCS / "catalogue"
SITE_IMAGES = DOCS / "assets" / "images"
SITE_IMAGES_URL = "/trinketa/assets/images"
STATISTICS = {locale: DOCS / f"statistiques.{locale}.md" for locale in ("fr", "en", "it")}


def text(value) -> str:
    """Convertit une cellule Excel en texte nettoyé, ou en chaîne vide."""
    return "" if value is None else str(value).strip()


def slug(value: str) -> str:
    """Produit un identifiant CSS simplifié à partir d'un libellé."""
    value = value.lower().strip()
    value = re.sub(r"[^a-z0-9-]+", "-", value)
    return value.strip("-")


def money(value) -> str:
    """Formate un montant en euros sans décimales inutiles."""
    if value in (None, ""):
        return ""
    try:
        return f"{float(value):,.0f} €".replace(",", " ")
    except (TypeError, ValueError):
        return text(value)


def number(value) -> float | None:
    """Convertit une valeur en nombre flottant ; renvoie ``None`` sinon."""
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def quantity(obj: dict) -> int:
    """Retourne la quantité physique du lot, avec 1 comme valeur de repli."""
    value = number(obj.get("Quantité"))
    return max(1, int(value)) if value is not None else 1


def unit_value(obj: dict, side: str) -> float | None:
    """Lit l'estimation unitaire basse ou haute d'un objet."""
    return number(obj.get(f"Estimation unitaire {side} (€)")) or number(obj.get(f"Estimation {side} (€)"))


def lot_value(obj: dict, side: str) -> float | None:
    """Retourne le total du lot, même si le moteur XLSX n'a pas mis en cache la formule."""
    total = number(obj.get(f"Estimation du lot {side} (€)"))
    if total is not None:
        return total
    unit = unit_value(obj, side)
    return unit * quantity(obj) if unit is not None else None


def estimate_range(low, high) -> str:
    """Présente une estimation basse/haute sous une forme lisible."""
    return " – ".join(v for v in [money(low), money(high)] if v) or "Non renseignée"


def treemap(data: Counter, empty_label: str = "Aucune donnée") -> str:
    """Génère les rectangles HTML d'une treemap proportionnelle aux effectifs."""
    if not data:
        return f'<p class="empty-state">{html.escape(empty_label)}</p>'
    total = sum(data.values())
    tiles = []
    for index, (label, count) in enumerate(sorted(data.items(), key=lambda item: (-item[1], item[0]))):
        share = 100 * count / total
        tiles.append(
            f'<div class="stat-treemap-tile stat-color-{index % 10}" '
            f'style="--weight:{count};--share:{share:.2f}%" '
            f'title="{html.escape(label, quote=True)} : {count} lot(s), {share:.1f} %">'
            f'<span>{html.escape(label)}</span><small>{share:.0f} %</small>'
            '</div>'
        )
    return '<div class="stat-treemap">' + "\n".join(tiles) + '</div>'


def top_with_other(data: Counter, limit: int = 12) -> Counter:
    """Conserve les catégories principales et regroupe les autres sous « Autres »."""
    ordered = data.most_common()
    if len(ordered) <= limit:
        return data
    result = Counter(dict(ordered[:limit]))
    result["Autres catégories"] = sum(count for _, count in ordered[limit:])
    return result


def location_category_histogram(furniture: list[dict], category_limit: int = 9) -> str:
    """Histogramme horizontal empilé : une barre par pièce, segmentée par catégorie."""
    category_totals = Counter(text(obj.get("Catégorie")) or "Non classé" for obj in furniture)
    main_categories = [label for label, _ in category_totals.most_common(category_limit)]
    legend_labels = main_categories + (["Autres catégories"] if len(category_totals) > category_limit else [])
    color_by_label = {label: index % 10 for index, label in enumerate(legend_labels)}

    matrix: dict[str, Counter] = defaultdict(Counter)
    for obj in furniture:
        location = text(obj.get("Localisation")) or "Localisation à préciser"
        category = text(obj.get("Catégorie")) or "Non classé"
        bucket = category if category in main_categories else "Autres catégories"
        matrix[location][bucket] += 1

    rows = []
    for location, counts in sorted(matrix.items(), key=lambda item: (-sum(item[1].values()), item[0])):
        total = sum(counts.values())
        segments = []
        for label in legend_labels:
            count = counts.get(label, 0)
            if not count:
                continue
            share = 100 * count / total
            visible_label = html.escape(str(count)) if share >= 7 else ""
            segments.append(
                f'<span class="stat-stack-segment stat-color-{color_by_label[label]}" '
                f'style="width:{share:.3f}%" title="{html.escape(label, quote=True)} : {count}" '
                f'aria-label="{html.escape(label, quote=True)} : {count}">{visible_label}</span>'
            )
        rows.append(
            '<div class="stat-stack-row">'
            f'<div class="stat-stack-label"><span>{html.escape(location)}</span><strong>{total}</strong></div>'
            f'<div class="stat-stack">{"".join(segments)}</div></div>'
        )

    legend = "".join(
        f'<span><i class="stat-color-{color_by_label[label]}"></i>{html.escape(label)}</span>'
        for label in legend_labels
    )
    return f'<div class="stat-stack-chart">{"".join(rows)}</div><div class="stat-legend">{legend}</div>'


def library_statistics_section(language: str) -> str:
    """Crée le conteneur des statistiques des livres dans la page commune.

    Les valeurs sont chargées côté navigateur depuis le JSON produit par
    ``generate_bibliotheque.py``.
    """
    labels = {
        "fr": {
            "title": "Bibliothèque",
            "filters": "Explorer le fonds",
            "period": "Période",
            "language": "Langue",
            "category": "Catégorie",
            "reset": "Réinitialiser",
            "map": "Carte des lieux d’édition",
            "cities": "Principaux lieux d’édition",
            "periods": "Répartition par période",
            "languages": "Répartition par langue",
            "categories": "Répartition par catégorie",
            "publishers": "Principaux éditeurs",
        },
        "en": {
            "title": "Library",
            "filters": "Explore the collection",
            "period": "Period",
            "language": "Language",
            "category": "Category",
            "reset": "Reset",
            "map": "Map of publishing places",
            "cities": "Main publishing places",
            "periods": "Distribution by period",
            "languages": "Distribution by language",
            "categories": "Distribution by category",
            "publishers": "Main publishers",
        },
        "it": {
            "title": "Biblioteca",
            "filters": "Esplora il fondo",
            "period": "Periodo",
            "language": "Lingua",
            "category": "Categoria",
            "reset": "Reimposta",
            "map": "Mappa dei luoghi di edizione",
            "cities": "Principali luoghi di edizione",
            "periods": "Distribuzione per periodo",
            "languages": "Distribuzione per lingua",
            "categories": "Distribuzione per categoria",
            "publishers": "Principali editori",
        },
    }[language]
    return f"""
<hr class="statistics-divider">

<div class="statistics-page library-statistics" data-library-statistics data-data-url="assets/data/bibliotheque-statistiques.json">
<header class="statistics-hero">
<h2>{labels["title"]}</h2>
</header>

<section class="library-filters" aria-label="{labels["filters"]}">
<label>{labels["period"]}<select data-filter="period"></select></label>
<label>{labels["language"]}<select data-filter="language"></select></label>
<label>{labels["category"]}<select data-filter="category"></select></label>
<button type="button" data-reset-filters>{labels["reset"]}</button>
</section>

<section class="library-kpis" data-kpis></section>

<div class="library-statistics-grid">
<section class="library-panel library-map-panel">
<h3>{labels["map"]}</h3>
<div class="library-map" data-library-map></div>
</section>
<section class="library-panel">
<h3>{labels["cities"]}</h3>
<div class="library-ranking" data-ranking="cities"></div>
</section>
<section class="library-panel">
<h3>{labels["periods"]}</h3>
<div class="library-chart" data-chart="periods"></div>
</section>
<section class="library-panel">
<h3>{labels["languages"]}</h3>
<div class="library-chart" data-chart="languages"></div>
</section>
<section class="library-panel">
<h3>{labels["categories"]}</h3>
<div class="library-chart" data-chart="categories"></div>
</section>
<section class="library-panel">
<h3>{labels["publishers"]}</h3>
<div class="library-ranking" data-ranking="publishers"></div>
</section>
</div>
</div>
"""


def generate_statistics(furniture: list[dict]) -> None:
    """Calcule et écrit la page Statistiques commune aux deux inventaires."""
    lows = [value for obj in furniture if (value := lot_value(obj, "basse")) is not None]
    highs = [value for obj in furniture if (value := lot_value(obj, "haute")) is not None]
    total_low = sum(lows)
    total_high = sum(highs)
    estimated_count = sum(
        1
        for obj in furniture
        if lot_value(obj, "basse") is not None
        or lot_value(obj, "haute") is not None
    )

    locations = Counter(text(obj.get("Localisation")) or "Localisation à préciser" for obj in furniture)
    categories = Counter(text(obj.get("Catégorie")) or "Non classé" for obj in furniture)
    location_chart = treemap(locations)
    category_chart = treemap(top_with_other(categories, limit=14))
    mixed_chart = location_category_histogram(furniture)
    coverage = (estimated_count / len(furniture) * 100) if furniture else 0
    item_count = sum(quantity(obj) for obj in furniture)

    translations = {
        "fr": {
            "title": "Statistiques", "subtitle": "Mobilier & objets d’art",
            "overview": "Vue d’ensemble", "estimate": "Estimation globale",
            "estimate_note": "Fourchette indicative cumulée", "lots": "Lots publiés",
            "items": "objet(s) inventorié(s)", "estimated": "Objets estimés",
            "catalogue": "du catalogue", "categories": "Catégories",
            "types": "Types renseignés", "reading": "Lecture des estimations",
            "notice": "Les montants sont des estimations documentaires indicatives. Leur addition donne un ordre de grandeur patrimonial, non une valeur de vente garantie.",
            "location": "Répartition par localisation", "category": "Répartition par catégorie",
            "mixed": "Catégories par localisation",
        },
        "en": {
            "title": "Statistics", "subtitle": "Furniture & works of art",
            "overview": "Overview", "estimate": "Total estimate",
            "estimate_note": "Cumulative indicative range", "lots": "Published lots",
            "items": "inventoried item(s)", "estimated": "Estimated objects",
            "catalogue": "of the catalogue", "categories": "Categories",
            "types": "Recorded types", "reading": "Reading the estimates",
            "notice": "Amounts are indicative documentary estimates. Their sum provides a broad heritage value, not a guaranteed sale price.",
            "location": "Distribution by location", "category": "Distribution by category",
            "mixed": "Categories by location",
        },
        "it": {
            "title": "Statistiche", "subtitle": "Arredi e oggetti d’arte",
            "overview": "Panoramica", "estimate": "Stima complessiva",
            "estimate_note": "Intervallo indicativo cumulativo", "lots": "Lotti pubblicati",
            "items": "oggetto/i inventariato/i", "estimated": "Oggetti stimati",
            "catalogue": "del catalogo", "categories": "Categorie",
            "types": "Tipi registrati", "reading": "Lettura delle stime",
            "notice": "Gli importi sono stime documentarie indicative. La loro somma fornisce un ordine di grandezza patrimoniale, non un prezzo di vendita garantito.",
            "location": "Distribuzione per ubicazione", "category": "Distribuzione per categoria",
            "mixed": "Categorie per ubicazione",
        },
    }

    for language, labels in translations.items():
        page = f"""# {labels["title"]}

## {labels["subtitle"]}

### {labels["overview"]}

<div class="stat-cards">
  <article class="stat-card stat-card-primary">
    <span>{labels["estimate"]}</span>
    <strong>{money(total_low)} – {money(total_high)}</strong>
    <small>{labels["estimate_note"]}</small>
  </article>
  <article class="stat-card">
    <span>{labels["lots"]}</span>
    <strong>{len(furniture)}</strong>
    <small>{item_count} {labels["items"]}</small>
  </article>
  <article class="stat-card">
    <span>{labels["estimated"]}</span>
    <strong>{estimated_count}</strong>
    <small>{coverage:.0f} % {labels["catalogue"]}</small>
  </article>
  <article class="stat-card">
    <span>{labels["categories"]}</span>
    <strong>{len(categories)}</strong>
    <small>{labels["types"]}</small>
  </article>
</div>

!!! note "{labels["reading"]}"
    {labels["notice"]}

### {labels["location"]}

<div class="stat-chart" role="img">
{location_chart}
</div>

### {labels["category"]}

<div class="stat-chart" role="img">
{category_chart}
</div>

### {labels["mixed"]}

<div class="stat-chart" role="img">
{mixed_chart}
</div>

{library_statistics_section(language)}
"""
        STATISTICS[language].write_text(page.rstrip() + "\n", encoding="utf-8")


def rows_as_dicts(sheet, header_row: int = 3):
    """Transforme chaque ligne Excel non vide en dictionnaire nommé par colonne."""
    headers = [text(cell.value) for cell in sheet[header_row]]
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        values = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        if any(value not in (None, "") for value in row):
            yield values


def require_columns(actual: set[str], expected: set[str], sheet_name: str):
    """Interrompt la génération si une colonne indispensable a été renommée."""
    missing = expected - actual
    if missing:
        raise ValueError(f"Feuille {sheet_name}: colonnes manquantes: {', '.join(sorted(missing))}")


def main() -> int:
    """Orchestre la lecture Excel, les images et la génération des pages."""
    if not WORKBOOK.exists():
        print(f"Classeur introuvable: {WORKBOOK}", file=sys.stderr)
        return 1

    wb = load_workbook(WORKBOOK, data_only=True)
    require_columns(
        {text(c.value) for c in wb["Mobilier"][3]},
        {"Identifiant", "Titre", "Catégorie", "Description", "Publié"},
        "Mobilier",
    )
    require_columns(
        {text(c.value) for c in wb["Photos"][3]},
        {"Identifiant objet", "Fichier", "Ordre", "Légende", "Image principale"},
        "Photos",
    )

    furniture = [r for r in rows_as_dicts(wb["Mobilier"]) if text(r["Publié"]).lower() in {"oui", "yes", "true", "1"}]
    photo_rows = list(rows_as_dicts(wb["Photos"]))
    photos_by_object = defaultdict(list)
    for photo in photo_rows:
        photos_by_object[text(photo["Identifiant objet"])].append(photo)
    for items in photos_by_object.values():
        items.sort(key=lambda p: (p["Ordre"] if isinstance(p["Ordre"], (int, float)) else 999, text(p["Fichier"])))

    INVENTORY_DIR.mkdir(parents=True, exist_ok=True)
    SITE_IMAGES.mkdir(parents=True, exist_ok=True)
    # Supprimer les anciennes sorties afin qu'une génération locale ne publie
    # pas simultanément les URL /catalogue/ et /inventaire/.
    if LEGACY_CATALOG_DIR.exists():
        for legacy in [*LEGACY_CATALOG_DIR.glob("MOB-*.md"), *LEGACY_CATALOG_DIR.glob("index*.md")]:
            legacy.unlink()
        try:
            LEGACY_CATALOG_DIR.rmdir()
        except OSError:
            # Conserver le dossier s'il contient un fichier qui n'appartient pas
            # au générateur, afin de ne jamais supprimer un contenu manuel.
            pass

    # Migration ponctuelle depuis l'ancienne structure non localisée.
    for legacy in [DOCS / "index.md", INVENTORY_DIR / "index.md", DOCS / "statistiques.md"]:
        if legacy.exists():
            legacy.unlink()
    for legacy in INVENTORY_DIR.glob("MOB-*.md"):
        if not any(legacy.name.endswith(f".{locale}.md") for locale in ("fr", "en", "it")):
            legacy.unlink()
    # Ne régénérer que les fiches françaises : les futures traductions .en.md
    # et .it.md doivent rester intactes.
    for old in INVENTORY_DIR.glob("MOB-*.fr.md"):
        old.unlink()

    cards = []
    categories = Counter()
    locations = Counter()
    warnings = []
    for obj in furniture:
        object_id = text(obj["Identifiant"])
        title = text(obj["Titre"])
        if not object_id or not title:
            warnings.append("Une ligne publiée sans identifiant ou sans titre a été ignorée.")
            continue

        photos = photos_by_object.get(object_id, [])
        available = []
        for photo in photos:
            raw_filename = text(photo["Fichier"]).strip()

            if not raw_filename:
                warnings.append(
                    f"{object_id}: nom de photographie vide ; ligne ignorée."
                )
                continue

            filename = Path(raw_filename).name
            source = SOURCE_PHOTOS / filename

            if source.is_file():
                shutil.copy2(source, SITE_IMAGES / filename)
                available.append(photo)
            elif source.is_dir():
                warnings.append(
                    f"{object_id}: le chemin photographique désigne un dossier : "
                    f"{raw_filename}"
                )
            else:
                warnings.append(
                    f"{object_id}: photographie introuvable : {filename}"
                )

        cover = next((p for p in available if text(p["Image principale"]).lower() in {"oui", "yes", "true", "1"}), available[0] if available else None)
        category = text(obj["Catégorie"]) or "Non classé"
        location = text(obj.get("Localisation")) or "Localisation à préciser"
        categories[category] += 1
        locations[location] += 1

        item_quantity = quantity(obj)
        quantity_label = f" · {item_quantity} items" if item_quantity > 1 else ""
        unit_estimate = estimate_range(unit_value(obj, "basse"), unit_value(obj, "haute"))
        lot_estimate = estimate_range(lot_value(obj, "basse"), lot_value(obj, "haute"))
        estimate = lot_estimate

        metadata = [
            ("Catégorie", category),
            ("Origine", text(obj.get("Origine"))),
            ("Matériaux", text(obj.get("Matériaux"))),
            ("Dimensions", text(obj.get("Dimensions"))),
            ("Localisation", text(obj.get("Localisation"))),
            ("Quantité", str(item_quantity)),
            ("Statut", text(obj.get("Statut"))),
        ]
        meta_html = "\n".join(
            f'<div class="record-field"><dt>{html.escape(label)}</dt><dd>{html.escape(value)}</dd></div>'
            for label, value in metadata if value
        )

        gallery = []
        for photo in available:
            filename = Path(text(photo["Fichier"])).name
            caption = text(photo["Légende"]) or title
            gallery.append(
                f'<figure><a href="{SITE_IMAGES_URL}/{html.escape(filename)}" target="_blank">'
                f'<img src="{SITE_IMAGES_URL}/{html.escape(filename)}" alt="{html.escape(caption)}" loading="lazy"></a>'
                f'<figcaption>{html.escape(caption)}</figcaption></figure>'
            )
        gallery_html = "\n".join(gallery) if gallery else '<p class="empty-state">Aucune photographie disponible.</p>'

        sources = []
        for key in ["Source 1", "Source 2"]:
            url = text(obj.get(key))
            if url:
                sources.append(f"- [{html.escape(url)}]({url})")
        source_block = "\n".join(sources) if sources else "_Aucune source renseignée._"

        association_ids = re.findall(r"MOB-\d{3}", text(obj.get("Associé à")), flags=re.IGNORECASE)
        association_block = " · ".join(
            f'<a href="../{associated.upper()}/">{html.escape(associated.upper())}</a>'
            for associated in association_ids
        ) or "_Aucun objet associé._"

        hero_visual = (
            f'<a href="{SITE_IMAGES_URL}/{html.escape(Path(text(cover["Fichier"])).name)}" target="_blank">'
            f'<img src="{SITE_IMAGES_URL}/{html.escape(Path(text(cover["Fichier"])).name)}" alt="{html.escape(title)}"></a>'
            if cover else '<div class="record-placeholder">Sans photographie</div>'
        )

        page = f"""<div class="lot-hero">
<div class="lot-visual">{hero_visual}</div>
<div class="lot-summary">
<p class="record-kicker">Lot {html.escape(object_id)} · {html.escape(category)}</p>
<h1>{html.escape(title)}</h1>
<p class="lot-dating">{html.escape(text(obj.get("Datation")) or "Datation à préciser")}</p>
<div class="lot-estimate"><span>Estimation du lot</span><strong>{html.escape(lot_estimate)}</strong></div>
<p class="lot-unit-estimate">Estimation unitaire : {html.escape(unit_estimate)}</p>

<dl class="record-grid">
{meta_html}
</dl>
</div>
</div>

## Description

{text(obj.get("Description")) or "_Description à compléter._"}

## État de conservation

{text(obj.get("État")) or "_État à compléter._"}

## Objets associés

{association_block}

## Photographies

<div class="gallery">
{gallery_html}
</div>

## Sources et comparaisons

{source_block}

<p class="notice">L’identification et l’estimation sont indicatives et peuvent évoluer avec de nouvelles mesures, photographies ou expertises.</p>
"""
        (INVENTORY_DIR / f"{object_id}.fr.md").write_text(page, encoding="utf-8")

        cover_html = (
            f'<img src="{SITE_IMAGES_URL}/{html.escape(Path(text(cover["Fichier"])).name)}" alt="{html.escape(title)}">'
            if cover else '<div class="card-placeholder">Sans photographie</div>'
        )
        cards.append(
            f'<article class="catalog-card" data-location="{html.escape(location, quote=True)}" '
            f'data-category="{html.escape(category, quote=True)}"><a href="{object_id}/">'
            f'{cover_html}'
            f'<div class="catalog-card-body"><div class="card-lot"><span>Lot {html.escape(object_id)}</span>'
            f'<span>{html.escape(category)}</span></div><h2>{html.escape(title)}</h2>'
            f'<p class="card-date">{html.escape(text(obj.get("Datation")) or "Datation à préciser")}'
            f'{html.escape(quantity_label)}</p>'
            f'<p class="card-estimate"><span>Estimation</span><strong>{html.escape(estimate)}</strong></p></div></a></article>'
        )

    location_options = "\n".join(
        f'<option value="{html.escape(label, quote=True)}">{html.escape(label)} ({count})</option>'
        for label, count in sorted(locations.items())
    )
    category_options = "\n".join(
        f'<option value="{html.escape(label, quote=True)}">{html.escape(label)} ({count})</option>'
        for label, count in sorted(categories.items())
    )
    catalogue = f"""<div class="catalogue-heading">
  <h1>Inventaire</h1>
</div>

<form class="catalog-filters" data-catalog-filters>
  <div class="catalog-filter catalog-search">
    <label for="catalog-query">Rechercher</label>
    <input id="catalog-query" type="search" name="q" placeholder="Titre, lot, époque…" autocomplete="off">
  </div>
  <div class="catalog-filter">
    <label for="catalog-location">Localisation</label>
    <select id="catalog-location" name="location">
      <option value="">Toutes les pièces</option>
      {location_options}
    </select>
  </div>
  <div class="catalog-filter">
    <label for="catalog-category">Type d’objet</label>
    <select id="catalog-category" name="category">
      <option value="">Tous les types</option>
      {category_options}
    </select>
  </div>
  <button class="catalog-reset" type="reset">Réinitialiser</button>
  <p class="catalog-result" aria-live="polite"><strong data-result-count>{len(cards)}</strong> lots affichés</p>
</form>

<div class="catalog-grid" data-catalog-grid>
{''.join(cards) if cards else '<p class="empty-state">Aucun objet publié.</p>'}
</div>
<p class="empty-state catalog-empty" data-catalog-empty hidden>Aucun lot ne correspond à ces critères.</p>
"""
    (INVENTORY_DIR / "index.fr.md").write_text(catalogue, encoding="utf-8")
    generate_statistics(furniture)

    if warnings:
        print("Avertissements:")
        for warning in warnings:
            print(f"- {warning}")
    print(f"Catalogue et statistiques générés: {len(cards)} objet(s), {sum(len(v) for v in photos_by_object.values())} photographie(s) référencée(s).")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
